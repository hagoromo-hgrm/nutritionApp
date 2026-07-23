#!/usr/bin/env python3
"""文部科学省の食品成分表Excelから、アプリ変換用CSVを作成する。

一般成分表と脂肪酸成分表を食品番号で結合する。Excelの記号は、未測定・
微量を空欄、括弧付きの推定値は数値として保持する。
"-"や"Tr"をゼロに置き換えないのは、データ不足を摂取ゼロと混同しないためである。
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

import openpyxl

from food_data_rules import clean_food_name, food_input_defaults, leading_category


CSV_COLUMNS = (
    "id", "official_name", "name", "maker", "barcode", "base_amount", "base_unit",
    "serving_amount", "serving_unit", "input_unit_conversions",
    "energy_kcal", "protein_g", "fat_g", "carbohydrate_g", "fiber_g", "salt_g",
    "calcium_mg", "iron_mg", "vitamin_a_mcg", "vitamin_e_mg", "vitamin_b1_mg",
    "vitamin_b2_mg", "vitamin_c_mg", "saturated_fat_g",
)


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text in {"-", "Tr", "(Tr)"}:
        return ""
    if text.startswith("(") and text.endswith(")"):
        text = text[1:-1].strip()
    return text


def food_id(value: Any) -> str:
    code = clean_value(value)
    return f"mext_{code.zfill(5)}"


def load_saturated_fat_values(input_xlsx: Path) -> dict[str, tuple[str, str]]:
    """脂肪酸成分表第1表から食品番号、食品名、飽和脂肪酸を読み取る。"""
    book = openpyxl.load_workbook(input_xlsx, read_only=True, data_only=True)
    try:
        sheet = book["表全体"]
        saturated_fat_column = None
        data_start = None
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if "FASAT" in row:
                saturated_fat_column = row.index("FASAT")
                data_start = row_number + 1
                break
        if saturated_fat_column is None or data_start is None:
            raise RuntimeError("脂肪酸成分表のFASAT列を特定できません。")

        values: dict[str, tuple[str, str]] = {}
        for row in sheet.iter_rows(min_row=data_start, values_only=True):
            if len(row) <= max(3, saturated_fat_column):
                continue
            raw_code = clean_value(row[1])
            official_name = clean_value(row[3])
            if not raw_code or not official_name:
                continue
            code = raw_code.zfill(5)
            if code in values:
                raise RuntimeError(f"脂肪酸成分表に食品番号の重複があります: {code}")
            values[code] = (official_name, clean_value(row[saturated_fat_column]))

        if not values:
            raise RuntimeError("脂肪酸成分表から食品データを読み取れません。")
        return values
    finally:
        book.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="MEXT食品成分表Excelを変換用CSVへ抽出")
    parser.add_argument("input_xlsx", type=Path)
    parser.add_argument("output_csv", type=Path)
    parser.add_argument("--fatty-acids-xlsx", type=Path, required=True)
    args = parser.parse_args()

    saturated_fat_values = load_saturated_fat_values(args.fatty_acids_xlsx)
    book = openpyxl.load_workbook(args.input_xlsx, read_only=True, data_only=True)
    sheet = book["表全体"]
    data_start = None
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if len(row) > 3 and clean_value(row[3]) == "成分識別子":
            data_start = row_number + 1
            break
    if data_start is None:
        raise RuntimeError("食品データの開始行を特定できません。")

    rows: list[dict[str, str]] = []
    matched_saturated_fat_ids: set[str] = set()
    try:
        for row in sheet.iter_rows(min_row=data_start, values_only=True):
            if len(row) < 61 or not clean_value(row[1]) or not clean_value(row[3]):
                continue
            code = clean_value(row[1]).zfill(5)
            raw_name = clean_value(row[3])
            saturated_fat_record = saturated_fat_values.get(code)
            if saturated_fat_record is not None:
                fatty_acid_name, saturated_fat = saturated_fat_record
                if fatty_acid_name != raw_name:
                    raise RuntimeError(
                        f"食品番号{code}の食品名が一般成分表と脂肪酸成分表で一致しません: "
                        f"{raw_name!r} != {fatty_acid_name!r}"
                    )
                matched_saturated_fat_ids.add(code)
            else:
                saturated_fat = ""

            defaults = food_input_defaults(raw_name, leading_category(raw_name))
            rows.append({
                "id": food_id(row[1]),
                "official_name": raw_name,
                "name": clean_food_name(raw_name),
                "maker": "",
                "barcode": "",
                "base_amount": "100",
                "base_unit": "g",
                "serving_amount": "" if defaults.serving_amount is None else str(defaults.serving_amount),
                "serving_unit": defaults.serving_unit or "",
                "input_unit_conversions": json.dumps(
                    [
                        {"unit": conversion.unit, "baseAmount": conversion.base_amount}
                        for conversion in defaults.input_unit_conversions
                    ],
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
                # MEXTの可食部100g当たりの値を変換・丸めずに保持する。
                "energy_kcal": clean_value(row[6]),
                "protein_g": clean_value(row[9]),
                "fat_g": clean_value(row[12]),
                "carbohydrate_g": clean_value(row[20]),
                "fiber_g": clean_value(row[18]),
                "salt_g": clean_value(row[60]),
                "calcium_mg": clean_value(row[25]),
                "iron_mg": clean_value(row[28]),
                "vitamin_a_mcg": clean_value(row[42]),
                "vitamin_e_mg": clean_value(row[44]),
                "vitamin_b1_mg": clean_value(row[49]),
                "vitamin_b2_mg": clean_value(row[50]),
                "vitamin_c_mg": clean_value(row[58]),
                "saturated_fat_g": clean_value(saturated_fat),
            })
    finally:
        book.close()

    unmatched_saturated_fat_ids = set(saturated_fat_values) - matched_saturated_fat_ids
    if unmatched_saturated_fat_ids:
        sample = ", ".join(sorted(unmatched_saturated_fat_ids)[:5])
        raise RuntimeError(f"一般成分表に存在しない脂肪酸成分表の食品番号があります: {sample}")

    args.output_csv.parent.mkdir(parents=True, exist_ok=True)
    with args.output_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    populated = sum(bool(row["saturated_fat_g"]) for row in rows)
    print(
        f"extracted foods={len(rows)} saturated_fat={populated} "
        f"output={args.output_csv}"
    )


if __name__ == "__main__":
    main()
